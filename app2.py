

""" weight_p=input( 'how much you weigh? ')
weight_k= 0.45* int(weight_p)
print(weight_k)
by=input('what is your birth year:')
#age=2024-by # error since int-str
age = 2024-int(by) # it coverts it to int
print(age) # it can not concatenate string with int
name=input('what is your name? ')
color=input('what color do you like? ')
print(name + ' likes ' + color)
course=''' how are you doing  
I wish you a happy life'''

sheet='hello test'

print(sheet[0])
print(sheet[0:5])# the 5 is not included
print(sheet[:])
first='john'
last='smith'
print(first + ' [' + last + '] '  +' is a coder')# the longer way
msg=f'{first} [{last}] is a coder' #formated string use {} to dynamically insert values to string
print(msg)
print(len(first))
#IF
price=1000000
is_good_credit=True
if is_good_credit:
    down_pay= 0.1 * price
else:
    down_pay=0.2 * price
print(f'${down_pay}')
#LOGIC
is_smart=True
is_girl=False
if is_smart and is_girl:
    print('I know her')
elif is_smart and not is_girl:
    print('I really do not know her')
#COMPARISION
# practice
name=input('What is your name:')
print('my name is :' + name)
length=len(name)
if length < 3:
    print('name must be at least 3 chars long')
elif length > 50:
    print('name can be a maximum of 50 chars')
else:
  print('name looks good!')
#practice for all
weight=input('weight:')# HERE WE CAN MAKE INT(INPUT)
choice=input('(L)bs or (k)g:')
if choice.lower() == 'l':
    lbs = int(weight) / 10
    print('you are' + f'{lbs}' +'lbs')
elif choice== 'k':
    kg = int(weight) - 50
    print( 'you are ' + f'{kg}' + 'kilo')# WE CAN SIMPLY USE A FORMATTED STRING f 'you are ...'
#WhileLoops
i=1
j=2
while i<=5:
    print('*' * i)
    print('*' * j)
    i=i+1
    j=j+2
print('done')
#Game 1
numb = 7
guess_count = 0
guess_limit = 3
while guess_count < guess_limit:
    guess = int(input('your guess: '))
    guess_count += 1
    if guess == numb:
        print('You won!')
        break
    elif guess < numb and guess_count < guess_limit:
        print('It\'s higher, try again')
    elif guess > numb and guess_count < guess_limit:
        print('It\'s lower, try again')
else:
    if guess != numb:
        print('You failed')
# Game 2
action=""
car_started=False
while True:
    action = input('>').lower().strip()
    if action == 'help':
        print(''' 
start - to start the car
stop - to stop the car
quit - to exit
        ''')
    elif action == 'start':
        if car_started:
           print('car already started')
        else:
            car_started=True
            print('starting the car....')
    elif action == 'stop':
        if not car_started:
           print('car already stopped')
        else:
            car_started=False
            print('stopping the car....')
    elif action == 'quit':
        print('Exiting the program.')
        break # to make it quit, we used the break
    else:
        print('I do not understand.')
#For loops
for item in ['Hindi', 'kubra']:
    print(item)
for items in range(6,10, 2): #here 2 i s step
   print(items)# 10 is not included
prices=[10,20,30]
total = 0
for price in prices:
    total += price
print(f 'total: {total}')
#Nested loop
for x in range(4):
    for y in range(3):
        print(f'({x},{y})')
# CHALLENGE
numbers=[5,2,5,2,2]
for num in numbers:
      print('*' * num )
# one solution, but with nested
numbers=[5,2,5,2,2]
for num in numbers:
    output=''
    for count in range (num):
      output += '*'
    print(output)
# print an L
numbers=[2,2,2,2,7]
for num in numbers:
    l = ''
    for count in range(num):
        l += '*'
    print(l)
#Lists
names=['kubra', 'sumru', 'sevilay']
names[0]='Kubra'
print(names[-1]) # the last
print(names[0:2])# they simply print the new list
from os.path import split

#practice finding the largest number in the list
basket = [20, 35, 10, 69]
maxi = basket[0]  # Initialize maxi with the first element
for numbers in basket:
    if numbers > maxi:
        maxi = numbers  # Update maxi if the current number is greater
print(f'This is the max: {maxi}')
#more questions on list
list1 = ["M", "na", "i", "Ke"]
list2 = ["y", "me", "s", "lly"]
list_n=[]
for i in range (len(list1)):#or we can use for i1
        combined= list1[i]+list2[i]
        list_n.append(combined)
print(list_n)
#square of number
numbers = [1, 2, 3, 4, 5, 6, 7]
list_square=[]
for i in range (len(numbers)):
    square=numbers[i]*numbers[i]
    list_square.append(square)
print(list_square)
#concatenate two lists
list3= ["Hello ", "take "]
list4 = ["Dear", "Sir"]
list_new=[]
for i, j in zip(list3,list4):
    combin=i + j
    list_new.append(combin)
print(list_new)
#or
res=[i+j for i in list3 for j in list4]
print(res)
#remove empty set
list1 = ["Mike", "", "Emma", "Kelly", "", "Brad"]
# remove None from list1 and convert result into list
res = list(filter(None, list1))
print(res)
#append
list1 = [10, 20, [300, 400, [5000, 6000], 500], 30, 40]

# understand indexing
# list1[0] = 10
# list1[1] = 20
# list1[2] = [300, 400, [5000, 6000], 500]
# list1[2][2] = [5000, 6000]
# list1[2][2][1] = 6000

# solution
list1[2][2].append(7000)
print(list1)
#extend
list1 = ["a", "b", ["c", ["d", "e", ["f", "g"], "k"], "l"], "m", "n"]
sub_list = ["h", "i", "j"]

# understand indexing
# list1[2] = ['c', ['d', 'e', ['f', 'g'], 'k'], 'l']
# list1[2][1] = ['d', 'e', ['f', 'g'], 'k']
# list1[2][1][2] = ['f', 'g']

# solution
list1[2][1][2].extend(sub_list)
print(list1)
#replace
list1 = [5, 10, 15, 20, 25, 50, 20]

# get the first occurrence index
index = list1.index(20)

# update item present at location
list1[index] = 200
print(list1)
#remove
list1 = [5, 20, 15, 20, 25, 50, 20]

# list comprehension
# remove specific items and return a new list
def remove_value(sample_list, val):
    return [i for i in sample_list if i != val]

res = remove_value(list1, 20)
print(res)
'''2d list
#[5,2,1,5,7,4]
#use append for last, and insert(where we want, what number added)
# index(5)..then it returns the index to find if  5 is in the list, also 50 in list boolean value
#count(5) which is 2, sort it sort it in place
#remove a duplicated from the list
numbers=[5,2,1,5,7,4]
x=[]
for item in numbers:
    if item not in x:
        x.append(item)
print(x)
# tuples are like list but with parenthesis and not edited only info wise
tuple1 = (11, 22)
tuple2 = (99, 88)
tuple1, tuple2 = tuple2, tuple1
print(tuple2)
print(tuple1)
# for copying tuple
tuple1 = (11, 22, 33, 44, 55, 66)
tuple2 = tuple1[3:-1]
print(tuple2)
# unpacking to retrieve form a tuple coordinate=(1,2,3) then x,y,z=coordinate print(y)
ph=input('Phone:')
number={'1': 'one',
'2': 'two',
'3':'three'
}
output=' '
for ch in ph:
    output+=number.get(ch,'!') + " "
print(output)
#Emoji convertor
mssg=input('>')
words=mssg.split(' ')
emoji={
    ':)': '😊',
    ':(' :'😒'
}
output_e=' '
for word in words:
    output_e+=emoji.get(word, word) + " "
print(output_e) 
#extract a function for the above
def sp(mssg):
    words=mssg.split(" ")
    emoji = {
        ':)': '😊',
        ':(': '😒'
    }
    change(words, emoji)
def change( words, emoji):
    output_e = ' '
    for word in words:
        output_e += emoji.get(word, word) + " "
    print(output_e)
sp('Hello :) how are you :(')
def em_con(mssg):
    words = mssg.split(' ')
    emoji = {
        ':)': '😊',
        ':(': '😒'
    }
    output_e = ' '
    for word in words:
      output_e += emoji.get(word, word) + " "
    return output_e
mssg=input('>')
res=em_con(mssg)
print(res)
# try and except
try:
    age=int(input("Age:"))
    print( age)
except ValueError:
    print(" age can not be other than integer")
#classes
# to define new types to model real concepts, collects every class
class SquareAction:
    def __init__(self, x, y):
        self.x=x
        self.y=y
    def move(self):
        print('move')
    def draw(self):
        print('draw')
S=SquareAction(10,20)
print(S.x)
S.draw()
#constructors
# a function that gets called at the time of calling
def __init__(self, x, y):
    self.x = x
    self.y = y
#Exercise
class Person:
    def __init__(self,name,age):
         self.name=name
         self.name2 = age
         self.talk_count=0

    def talk (self):
        self.talk_count+=1
        print(f'{self.name} talks. This is the {self.talk_count} time(s).')
    def greet(self):
        print(f'I am {self.name}')

    def reset_talk_count(self):
        self.talk_count = 0  # Reset the counter to 0
        print(f'{self.name}\'s talk count has been reset.')
name=input(" what is your name:")
age = input("How old are you: ")
P = Person(name,age)
P.greet()

# The person talks a few times
P.talk()
P.talk()

# Reset the talk count
P.reset_talk_count()

# Person talks again
P.talk()

name=input(" what is your name:")
age = input("How old are you: ")
Z=Person(name,age)

Z.greet()

# The person talks a few times
Z.talk()
Z.talk()

# Reset the talk count
Z.reset_talk_count()

# Person talks again
Z.talk()
class Person:
    def __init__(self, name, age):
        self.name = name
        self.age = age
        self.talk_count = 0

    def talk(self):
        self.talk_count += 1
        print(f'{self.name} talks. This is the {self.talk_count} time(s).')

    def greet(self):
        print(f'Hello, my name is {self.name} and I am {self.age} years old.')

    def reset_talk_count(self):
        self.talk_count = 0  # Reset the counter to 0
        print(f'{self.name}\'s talk count has been reset.')

# Loop to keep interacting with the program
while True:
    choice = input("Choose an option: 1) Create Person 2) Talk 3) Reset Talk Count 4) Exit\n> ")

    if choice == '1':
        # Create a new person
        name = input("What is your name: ")
        age = input("How old are you: ")
        P = Person(name, age)
        P.greet()

    elif choice == '2':
        # Person talks
        if 'P' in locals():  # Check if a person has been created
            P.talk()
        else:
            print("Please create a person first.")

    elif choice == '3':
        # Reset talk count
        if 'P' in locals():
            P.reset_talk_count()
        else:
            print("Please create a person first.")

    elif choice == '4':
        print("Exiting the program.")
        break  # Exit the loop

    else:
        print("Invalid choice, please choose again.")
#inheritance
class Mammal:
    def walk(self):
        print ('walk')
class Dog(Mammal):
    def bark (self): # we can also write pass
        print('bark')
class Cat(Mammal):
    def walk(self):
        pass
d=Dog()
d.walk()
d.bark()
#Modules
# organize file into multiple files
print (Conv.to_kg(20))
# maximum number
import Conv
numbers=[106,35,63,42,98,65,89,23,45,73,89]
F = Conv.FindMax(numbers)
print(max(numbers))
#packages
# organize related modcules in to packages( mall{ men: shoes women:cloth
import ecommerce.shipping
ecommerce.shipping.cal_ship()# is very long
from ecommerce.shipping import cal_ship
cal_ship()
#Built-in modules
import random
members=['john','mary','siri']
leader=random.choice(members)
print(leader)
for i in range(3):
    print(random.randint(20,30))
    print(random.random())
#Exercise
import random
class Dice:
    def __init__(self,tuples_range):
        self.tuples_range=tuples_range
    def roll(self):
        selected= random.choice(self.tuples_range)
        return selected
tuple_range=[(1,2),(3,4),(3,5)]# or we can simply generate a randint form 1 t0 6 (1,6)
d=Dice(tuple_range)
print( d.roll() )
#short answer
class Dice:
    def roll(self):
        first=random.randint(1,6)
        second=random.randint(1, 6)
        return first, second
d=Dice()
print(d.roll())
# how to work on directories
from pathlib import Path
p = Path ("emailsworks,(works)")
print(p.exists())
print(p.rename('emailsworks,(works)'))
p.rmdir()
#use glob to find files...it is another level
# Yay Project One, automate the process and also add a graph
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def process_workbook(filename):
    wb = xl.load_workbook('filename')
    sheet = wb['Sheet1']
    # print(cell.value)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = 0.9 * cell.value
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # to add a chart
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    B = BarChart()
    B.add_data(values)
    sheet.add_chart(B, 'e2')
    wb.save('filename')

process_workbook(tansactions2.xlsx)"""
#Machine Learning
'''subset of ai,scan an image that tells if dog is dog, and cat is cat
  give it a lots of data,then train it, for example in self-deriving cars, robotics 
  Steps:
  1.import data
  2.clean the data(so that it does not learn bad pattern
  3.split the data into training and test
  4.create a model
  5. train model,
  6.make prediction, 
  7.evaluate and improve
  # Libraries.Numpy,pandas,MatPlotLib,scikit-learn'''

